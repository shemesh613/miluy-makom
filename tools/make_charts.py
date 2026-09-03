# -*- coding: utf-8 -*-
"""בונה שני דפי A3 מקובץ המערכת: אחד לפי כיתות ואחד לפי מלמדים."""
import io, sys, re, json, html
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
XLSX = 'maarechet_14-8.xlsx'
DAYS = ['ראשון', 'שני', 'שלישי', 'רביעי', 'חמישי', 'שישי']


def sheet_rows(ws):
    return [['' if c is None else str(c).strip() for c in r]
            for r in ws.iter_rows(values_only=True)]


def blocks(rows):
    """כל בלוק = כותרת (כיתה/מלמד) + טבלת שעות × ימים. איתור דינמי."""
    heads = [i for i, r in enumerate(rows) if any(c == 'ראשון' for c in r)]
    out = []
    for si, hr in enumerate(heads):
        end = heads[si + 1] - 1 if si + 1 < len(heads) else len(rows)
        for c, cell in enumerate(rows[hr]):
            if cell != 'ראשון':
                continue
            prev = rows[hr - 1] if hr else []
            title = re.sub(r'\s+', ' ', (prev[c - 1] if c - 1 < len(prev) else '')).strip()
            if not title:
                continue
            grid = []
            for i in range(hr + 1, end):
                row = rows[i] if i < len(rows) else []
                label = (row[c - 1] if c - 1 < len(row) else '').strip()
                if not label:
                    continue
                vals = [(row[c + 1 + d] if c + 1 + d < len(row) else '') for d in range(6)]
                if not any(vals) and not label.replace('.', '').isdigit():
                    continue
                grid.append((label, vals))
            while grid and not any(v.strip() for v in grid[-1][1]):
                grid.pop()                      # שורות זנב ריקות
            if grid:
                out.append({'title': title, 'grid': grid})
    return out


wb = openpyxl.load_workbook(XLSX, data_only=True)
by_class = blocks(sheet_rows(wb[next(n for n in wb.sheetnames if 'כתות' in n or 'כיתות' in n)]))
by_teacher = blocks(sheet_rows(wb[next(n for n in wb.sheetnames if 'מלמדים' in n)]))

# מיון הכיתות בסדר טבעי א1..ח
ORDER = {c: i for i, c in enumerate(
    ['א1', 'א2', 'ב1', 'ב2', 'ג1', 'ג2', 'ד', 'ה1', 'ה2', 'ו1', 'ו2', 'ז', 'ח'])}
by_class.sort(key=lambda b: ORDER.get(b['title'].split()[0], 99))
by_teacher.sort(key=lambda b: b['title'])

CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { direction: rtl; font-family: "Segoe UI", "Arial Hebrew", Arial, sans-serif;
       background: #fff; color: #14181d; padding: 14px 20px 8px; }
.head { display: flex; align-items: baseline; gap: 16px; border-bottom: 3px solid #1d3557;
        padding-bottom: 7px; margin-bottom: 12px; }
.head h1 { font-size: 34px; color: #1d3557; letter-spacing: -0.5px; }
.head .meta { color: #6b7280; font-size: 19px; margin-inline-start: auto; }
.grid { column-gap: 9px; }
.box { display: inline-block; width: 100%; margin-bottom: 9px; }
.box { border: 1.5px solid #c9d2de; border-radius: 9px; overflow: hidden; break-inside: avoid; }
.box > h2 { background: #1d3557; color: #fff; font-size: 18px; padding: 4px 8px;
            text-align: center; font-weight: 700; }
table { width: 100%; border-collapse: collapse; table-layout: fixed; }
th, td { border: 1px solid #dbe2ea; padding: 2.5px 2px; text-align: center;
         font-size: 14px; line-height: 1.25; white-space: nowrap; overflow: hidden; }
th { background: #eef2f7; color: #33415c; font-weight: 700; font-size: 13px; }
td.h { background: #f6f8fb; color: #5a6472; font-weight: 700; width: 30px; font-size: 12.5px; }
tbody tr:nth-child(even) td:not(.h) { background: #fafbfd; }
td.e { color: #c3cad4; }
td.m { font-size: 12px; }      /* טקסט בינוני */
td.s { font-size: 10.5px; }    /* טקסט ארוך — מתכווץ במקום לגלוש */
td.xs { font-size: 9px; letter-spacing: -0.3px; }
td.n { color: #9b5d00; background: #fff8ea !important; font-size: 12.5px; }
.foot { margin-top: 6px; color: #8b93a0; font-size: 13px; text-align: center; }
"""

NONCLASS = ('הדרכה', 'גנים', 'קב"ש', 'ס. ערב', 'ס. משנה', 'סדר ערב', 'סדר משנה',
            'י. מחנכים', 'מחנך', 'תפילה')


def cell(v):
    raw = v
    v = html.escape(v)
    if not v:
        return '<td class="e">·</td>'
    cls = ['n'] if any(raw.startswith(p) for p in NONCLASS) else []
    n = len(raw)                       # תא ארוך מתכווץ במקום לדרוס את השכן
    if n > 10:
        cls.append('xs')
    elif n > 7:
        cls.append('s')
    elif n > 5:
        cls.append('m')
    attr = f' class="{" ".join(cls)}"' if cls else ''
    return f'<td{attr}>{v}</td>'


def render(title, meta, items, cols, out):
    boxes = []
    for b in items:
        rows = ''.join(
            f'<tr><td class="h">{html.escape(lbl)}</td>' + ''.join(cell(v) for v in vals) + '</tr>'
            for lbl, vals in b['grid'])
        boxes.append(
            f'<div class="box"><h2>{html.escape(b["title"])}</h2><table>'
            f'<thead><tr><th></th>' + ''.join(f'<th>{d}</th>' for d in DAYS) + '</tr></thead>'
            f'<tbody>{rows}</tbody></table></div>')
    doc = (f'<!doctype html><html lang="he" dir="rtl"><head><meta charset="utf-8">'
           f'<title>{title}</title><style>{CSS}</style></head><body>'
           f'<div class="head"><h1>{title}</h1><div class="meta">{meta}</div></div>'
           f'<div class="grid" style="column-count:{cols}">'
           + ''.join(boxes) +
           '</div><div class="foot">תלמוד תורה מוריה · מערכת תשפ"ז · עדכון 14.8.26 · '
           'משבצות בכתום = לא שיעור עם כיתה</div></body></html>')
    open(out, 'w', encoding='utf-8').write(doc)
    print('  %s → %d טבלאות' % (out, len(items)))


print('נבנה:')
render('מערכת תשפ"ז — לפי כיתות', '13 כיתות', by_class, 4, 'chart_classes.html')

# 29 מלמדים לא נכנסים לעמוד אחד בלי לחתוך שמות מקצועות — מפוצל לשניים
half = (len(by_teacher) + 1) // 2
for i, part in enumerate([by_teacher[:half], by_teacher[half:]], 1):
    render('מערכת תשפ"ז — לפי מלמדים (%d מתוך 2)' % i,
           '%d מלמדים' % len(part), part, 4, 'chart_teachers_%d.html' % i)
