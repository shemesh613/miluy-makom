# -*- coding: utf-8 -*-
"""Replicates miluy-makom's parseScheduleRows / parseMechanechRows in Python,
so the תשפ"ז timetable can be baked into index.html as the built-in schedule."""
import io, sys, re, json
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = r"maarechet_14-8.xlsx"   # מערכת תשפ"ז, עדכון 14.8.26 (מחליף את זו מ-6.8)

ALIASES = {'הרב שימשון': 'הרב שמשון'}
CLASS_TOKEN = re.compile(r'^[א-ח][12]?$')
HOUR_ORDER = ['בוקר', '4', '5', '6', '7', '8', '9', '10', 'סדר ערב']


def norm(n):
    t = str(n).strip()
    return ALIASES.get(t, t)


def extract_class(cell):
    for tok in re.split(r'[\s,/|]+', str(cell)):
        if CLASS_TOKEN.match(tok):
            return tok
    return None


def sheet_rows(ws):
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(['' if c is None else str(c).strip() for c in r])
    return rows


# פעילויות שאינן שיעור עם כיתה. השעה **תפוסה** — המלמד לא פנוי למלא מקום —
# אבל אסור לשייך לה כיתה: אין כיתה שמחכה, ולכן גם אם הוא נעדר אין מה לכסות.
# ההיגיון כאן הפוך מזה של תורנות חצר: שם "הדרכה" אינה חוסמת (כלל קבוצת הגיל
# קיים רק בגלל איפה הילדים נמצאים, ובהדרכה אין ילדים) — כאן היא כן חוסמת.
NON_CLASS_PREFIXES = (
    'י. מחנכים', 'ישיבת מחנכים', 'י.מחנכים',   # ישיבת צוות
    'הדרכה',                                    # הדרכת מלמדים
    'גנים',                                     # ליווי הגנים (חדש ב-14.8)
    'קב"ש',                                     # קבלת שבת — גם "קב\"ש ב2"
    'ס. משנה', 'סדר משנה', 'ס. ערב', 'סדר ערב',  # סדרים
)


def is_non_class(cell):
    c = str(cell).strip()
    return any(c.startswith(p) for p in NON_CLASS_PREFIXES)


def parse_schedule(rows):
    header_rows = [i for i, r in enumerate(rows) if any(c == 'ראשון' for c in r)]
    if not header_rows:
        raise SystemExit('לא נמצאה שורת ימים')
    acc = {}
    for si, hr in enumerate(header_rows):
        blocks = []
        for c, cell in enumerate(rows[hr]):
            if cell == 'ראשון':
                prev = rows[hr - 1] if hr - 1 >= 0 else []
                name = norm(prev[c - 1] if c - 1 < len(prev) else '')
                if name:
                    blocks.append((c - 1, name))
        end = header_rows[si + 1] - 1 if si + 1 < len(header_rows) else len(rows)
        for col, name in blocks:
            a = acc.setdefault(name, {'dayHours': {}, 'hourClasses': {}, 'meetings': set(), 'nonclass': set()})
            for i in range(hr + 1, end):
                row = rows[i] if i < len(rows) else []
                raw = row[col] if col < len(row) else ''
                try:
                    hour_num = int(float(raw))
                except (ValueError, TypeError):
                    continue
                if hour_num < 1 or hour_num > 12:
                    continue
                for d in range(6):
                    idx = col + 1 + d
                    cell = row[idx] if idx < len(row) else ''
                    if not cell:
                        continue
                    label = 'בוקר' if hour_num <= 3 else (str(hour_num) if hour_num <= 10 else None)
                    if label is None:
                        continue
                    a['dayHours'].setdefault(d, set()).add(label)
                    if is_non_class(cell):
                        a['nonclass'].add((d, label))
                        if str(cell).strip().startswith(('י. מחנכים', 'ישיבת מחנכים', 'י.מחנכים')):
                            a['meetings'].add((d, label))
                        continue
                    cls = extract_class(cell)
                    if cls:
                        hc = a['hourClasses'].setdefault(d, {})
                        hc.setdefault(label, cls)
    out = {}
    for name, data in acc.items():
        days = sorted(data['dayHours'])
        if not days:
            continue
        day_hours, day_classes = {}, {}
        for d in days:
            day_hours[d] = sorted(data['dayHours'][d], key=HOUR_ORDER.index)
            bases = {c.replace('1', '').replace('2', '') for c in data['hourClasses'].get(d, {}).values()}
            day_classes[d] = sorted(bases)
        out[name] = {'days': days, 'dayHours': day_hours, 'dayClasses': day_classes,
                     'hourClasses': data['hourClasses'],
                     'meetings': sorted('%s|%s' % m for m in data['meetings']),
                     'nonclass': sorted('%s|%s' % m for m in data['nonclass'])}
    return out


def resolve(raw, known):
    if raw in known:
        return raw
    m = [k for k in known if raw in k or k in raw]
    return m[0] if len(m) == 1 else raw


def parse_mechanech(rows, known):
    mech = {}
    for i, r in enumerate(rows):
        if not any(c == 'ראשון' for c in r):
            continue
        for c, cell in enumerate(r):
            if cell == 'ראשון':
                prev = rows[i - 1] if i - 1 >= 0 else []
                header = (prev[c - 1] if c - 1 < len(prev) else '').strip()
                m = re.match(r'^([א-ח][12]?)\s+(.+)$', header)
                if m:
                    mech[m.group(1)] = resolve(norm(m.group(2)), known)
    return mech


wb = openpyxl.load_workbook(XLSX, data_only=True)
print('גיליונות:', wb.sheetnames)
t_sheet = next((n for n in wb.sheetnames if 'מלמדים' in n), wb.sheetnames[-1])
c_sheet = next((n for n in wb.sheetnames if 'כתות' in n or 'כיתות' in n), None)

t_rows = sheet_rows(wb[t_sheet])
parsed = parse_schedule(t_rows)
mech = parse_mechanech(sheet_rows(wb[c_sheet]), list(parsed)) if c_sheet else {}
mech['ו2'] = 'הרב משה'          # בקובץ רשום "משה" בלבד — התאמה חד-משמעית

# מחנך שהתא שלו מכיל רק את שם המקצוע ("ב\"מ", "ברכות", "מחנך", "קב\"ש") מלמד
# את כיתתו שלו. בלי ההשלמה הזו מחנכי ז/ח יוצאים בלי אף כיתה.
own = {t: c for c, t in mech.items()}
filled = 0
for name, t in parsed.items():
    cls = own.get(name)
    if not cls:
        continue
    meets = set(t['nonclass'])
    for d in t['days']:
        hc = t['hourClasses'].setdefault(d, {})
        for h in t['dayHours'][d]:
            if h not in hc and '%s|%s' % (d, h) not in meets:
                hc[h] = cls
                filled += 1
        t['dayClasses'][d] = sorted({c.replace('1', '').replace('2', '')
                                     for c in hc.values()})
print('הושלמו %d משבצות של מחנכים לפי כיתתם' % filled)

# ---------------------------------------------------------------------------
# חסימות אישיות: מלמד שאינו בבית הספר בשעות מסוימות, גם כשבקובץ הן נראות
# כשעת חלון. בלי זה הוא מוצע כממלא מקום דווקא כשהוא לא בבניין.
# החסימה מוסיפה "תפוס" בלבד — היא לא מוחקת שיעורים שהוא כן מלמד אז.
# ---------------------------------------------------------------------------
PERSONAL_BLOCKS = {
    'הרב ינון': {'hours': ['4', '5'],
                 'reason': 'יוצא מהמתחם 10:05–12:30'},
    # מגיע ב-10:05 רק לצורך תורנות החצר בהפסקה; לפני כן אינו בבניין.
    # בשעה 4 עצמה הוא מלמד נביא ו1 — אין כאן זמינות חדשה למילוי מקום.
    'הרב פורת': {'hours': ['בוקר'], 'days': [0, 3],
                 'reason': 'אינו בבניין לפני 10:05'},
}

blocked = 0
for name, rule in PERSONAL_BLOCKS.items():
    t = parsed.get(name)
    if not t:
        print('⚠️ חסימה אישית ל%s — לא נמצא בקובץ' % name)
        continue
    days = rule.get('days') or list(range(6))
    for d in days:
        if d not in t['days']:
            continue                      # יום שהוא בכלל לא עובד בו
        hours = t['dayHours'].setdefault(d, [])
        for h in rule['hours']:
            if h not in hours:
                hours.append(h)
                blocked += 1
        hours.sort(key=HOUR_ORDER.index)
    print('חסימה אישית: %s — שעות %s (%s)'
          % (name, ', '.join(rule['hours']), rule['reason']))
print('  נוספו %d שעות חסומות' % blocked)

# ---------------------------------------------------------------------------
# ישיבת מחנכים: כשהיא מתקיימת — *כל* המחנכים תפוסים, גם אם בקובץ נשכח לציין
# את זה אצל מישהו. בלי הכלל הזה מחנך שהתא שלו ריק ייראה פנוי למילוי מקום.
# המחנכים = מי שמלמד בשעות 1–3 ברוב הימים (יום חופשי אחד לא פוסל) — וזו בדיוק
# רשימת 13 המחנכים מגיליון "לפי כתות".
# ---------------------------------------------------------------------------
from collections import Counter
slots = Counter(m for t in parsed.values() for m in t['meetings'])
if slots:
    meeting, seen = slots.most_common(1)[0]
    md, mh = meeting.split('|')
    md = int(md)
    added = []
    for name in own:                       # own = מחנך -> כיתה
        t = parsed.get(name)
        if not t:
            continue
        hours = t['dayHours'].setdefault(md, [])
        if mh not in hours:
            hours.append(mh)
            hours.sort(key=HOUR_ORDER.index)
            added.append(name)
        if md not in t['days']:
            t['days'] = sorted(t['days'] + [md])
    others = [m for m, _ in slots.items() if m != meeting]
    print('ישיבת מחנכים: יום %d שעה %s — %d/%d מחנכים סומנו בקובץ%s'
          % (md, mh, seen, len(own), (', נוספו: ' + ', '.join(added)) if added else ''))
    if others:
        print('  ⚠️ רשומות ישיבה בשעות אחרות: %s' % ', '.join(
            '%s (%s)' % (o, ', '.join(n for n, t in parsed.items() if o in t['meetings']))
            for o in others))
else:
    print('⚠️ לא נמצאה ישיבת מחנכים בקובץ')

total = sum(len(h) for t in parsed.values() for h in t['dayHours'].values())
print('גיליון מלמדים: %s | גיליון כיתות: %s' % (t_sheet, c_sheet))
print('מורים: %d | משבצות שבועיות: %d | מחנכים: %d' % (len(parsed), total, len(mech)))
print()
for n in sorted(parsed, key=lambda x: -sum(len(h) for h in parsed[x]['dayHours'].values())):
    print('  %-24s %2d ש"ש  ימים %s' % (
        n, sum(len(h) for h in parsed[n]['dayHours'].values()), parsed[n]['days']))
print()
print('מחנכים:', json.dumps(mech, ensure_ascii=False))

json.dump({'teachers': parsed, 'mechanech': mech},
          open('tashpaz_parsed.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
